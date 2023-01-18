import calendar
import time
from copy import deepcopy
from datetime import datetime

import openpyxl
import youtube_dl
from course_video.forms import AddCourseForm
from course_video.models import *
from django.contrib import messages
from django.core.paginator import EmptyPage
from django.core.paginator import PageNotAnInteger
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import HttpResponse
from django.shortcuts import redirect
from django.shortcuts import render
from django.views import View
from django.views.generic.edit import FormView
from django.views.generic.list import ListView
from members.models import Profile
from taggit.models import Tag


class Home(View):
    """
    Home class is used to display home page of an app.
    """

    def get(self, request):
        return render(request, "course/home.html")


class Main(ListView):
    """
    Main class is used to display the course,
    and progress bar status too, and after successfully adition of course email
    notification is sent.
    """

    model = Course
    paginate_by = 4
    template_name = "course/index.html"

    def get_context_data(self, **kwargs):
        context = super(Main, self).get_context_data(**kwargs)
        paginator = Paginator(Course.objects.select_related().all(), self.paginate_by)
        page = self.request.GET.get("page")
        try:
            course = paginator.page(page)
        except PageNotAnInteger:
            course = paginator.page(1)
        except EmptyPage:
            course = paginator.page(paginator.num_pages)

        course_id = []
        course = Course.objects.all()
        profile = Profile.objects.get(user=self.request.user)
        for item in course.iterator():
            course_list = Course.objects.get(id=item.id)
            course_id.append(course_list)
        progress = {}
        for data in course_id:
            global_count = 0
            list_item = PlaylistItem.objects.filter(playlist=data.id).filter(
                author=data.author
            )
            for title in list_item.iterator():
                if title.status == "Completed":
                    global_count = global_count + 1
                if global_count == 0:
                    percentage = 0
                else:
                    total = global_count
                    percentage = 100 * float(total) / float(list_item.count())
                progress[data.id] = round(percentage)
            self.request.session["progress"] = progress
        print(progress)
        context["course"] = paginator.page(context["page_obj"].number)
        context["per"] = progress
        context["profile"] = profile
        return context


class AddCourse(FormView):
    """
    AddCourse class is used to add course to list_of_playlist
    """

    template_name = "course/add_course.html"
    form_class = AddCourseForm
    success_url = "course"

    def form_valid(self, form):

        link = form.cleaned_data["link"]
        tag = form.cleaned_data["tags"]
        ydl = youtube_dl.YoutubeDL({"ignoreerrors": True, "quiet": True})
        with ydl:
            result = ydl.extract_info(
                link, download=False  # We just want to extract the info
            )
        playlist_info = []
        for video in result["entries"]:
            video_url = dict(
                (k, video[k])
                for k in ["title", "duration", "webpage_url"]
                if k in video
            )
            playlist_info.append(video_url.copy())

        form.instance.title = result.get("title")

        form.instance.author = self.request.user
        obj = form.save()
        obj.tags.add(*tag)
        cource = Course.objects.get(id=obj.id)
        for i in playlist_info:
            title = i.get("title")
            duration = i.get("duration")
            playlist_url = i.get("webpage_url")
            list_store = PlaylistItem(
                list_item=title,
                time=duration,
                link=playlist_url,
                playlist=cource,
                author=self.request.user,
            )
            list_store.save()
        messages.success(
            self.request, "Sent Email Successfully...Check your mail please"
        )
        return super().form_valid(form)


class TagListing(View):
    """
    TagListing class is used to filter out course based on given tags

    Args:-
        tags = slug of tag name
    return-
        list of course based on provided tags.
    """

    def get(self, request, tags):
        course = Course.objects.all()
        tags = Tag.objects.filter(slug=tags).values_list("name", flat=True)
        posts = course.filter(tags__name__in=tags)
        info = {
            "tags": posts,
        }
        return render(request, "course/tag.html", info)


class Listing(View):
    """
    Listing class is used to open particular video playlist and maintain the progress
    bar status.

    Args-
     id - This is the id of particular video playlist.

    return-
        list of video playlist.
    """

    def get(self, request, id):
        status_data = ["Yet to Start", "In Progress", "On Hold", "Completed"]
        course_list = Course.objects.get(id=id)
        list_item = PlaylistItem.objects.filter(playlist=course_list.id)
        request.session["id"] = id

        course_info = {
            "videoDetail": list_item,
            "id": id,
            "overall": status_data,
            "title": course_list.title,
        }

        return render(request, "course/List.html", course_info)


class Search(View):
    """
    Search class is used to filter out course which is written on search box.
    """

    def get(self, request):
        query = self.request.GET["query"]
        allCourse = Course.objects.filter(title__icontains=query)
        params = {"allCourse": allCourse}
        return render(request, "course/search.html", params)


class Editing(View):
    """
    Editing class is used to edit the title of course.
    """

    def post(self, request, id):
        course = Course.objects.get(id=id)
        title = request.POST["title"]
        course.title = title
        course.save(update_fields=["title"])
        PlaylistItem.objects.filter(playlist=course.id).update(playlist=course.id)
        messages.success(request, "Successfully updated title")
        return redirect("course_video:course")

    def get(self, request, id):
        course = Course.objects.get(id=id)
        return render(request, "course/edit.html", {"course": course, "id": id})


class GetStatus(View):
    """
    GetStatus class is used to store current status value
    ("Yet to Start", "In Progress", "On Hold", "Completed")
    to DB.
    """

    def post(self, request, id):
        update_status = PlaylistItem.objects.get(id=id)
        list_id = request.session.get("id")
        choice = request.POST["choice"]
        update_status.author = self.request.user
        update_status.status = choice
        update_status.save(update_fields=["status", "author"])

        return redirect("course_video:listing", id=list_id)


class Clone(View):
    """
    Clone class is used to make clone of public course of other user.
    """

    def get(self, request, id):
        course = Course.objects.get(id=id)
        if self.request.user.id != course.author.id and course.public:
            old_obj = deepcopy(course)
            playlist = PlaylistItem.objects.filter(playlist=old_obj.id)
            old_playlist = deepcopy(playlist)

            old_tag = old_obj.tags.all()
            course.id = None
            course.author = self.request.user

            course.save()
            for tag in old_tag.iterator():
                course.tags.add(tag.name)
            for item in old_playlist:
                PlaylistItem.objects.create(
                    list_item=item.list_item,
                    time=item.time,
                    playlist=course,
                    link=item.link,
                    status="Yet To Start",
                    author=self.request.user,
                )
        return redirect("main")


class CourseCsv(View):
    """
    CourseCsv class is used to download private course in excel format.
    """

    def get(self, request):
        wb = openpyxl.Workbook()
        sheetOne = wb.create_sheet("ListOfPlaylist")
        course = Course.objects.filter(public=False)
        data_view = []

        for item in course.iterator():
            if request.user.id == item.author.id:
                progress = request.session.get("progress")
                for i in progress.keys():
                    if i.find(str(item.id)) > -1:
                        progress_data = f"{progress[i]} %"
                old_tag = item.tags.all()
                for tag in old_tag.iterator():
                    data = (item.title, item.link, tag.name, progress_data)

                data_view.append(list(data))
                item_list = [tuple(l) for l in data_view]
                add_tuple = ("Title", "Link", "Tag", "Progress")
                item_list.insert(0, add_tuple)

        for playlist_item in item_list:
            sheetOne.append(playlist_item)
        wb.remove(wb["Sheet"])
        current_GMT = time.gmtime()

        # ts stores timestamp
        ts = calendar.timegm(current_GMT)
        date_time = datetime.fromtimestamp(ts)

        # convert timestamp to string in dd-mm-yyyy HH:MM:SS
        str_date_time = date_time.strftime("%d-%m-%Y, %H:%M:%S")
        wb.save(f"{request.user} {str_date_time}.xlsx")
        return redirect("main")


class Upload(View):
    """
    Upload class is used to added course by just providing excel sheets in proper format.
    """

    def post(self, request):
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        sheets = wb.sheetnames
        excel_data = []
        excel_playlist = list()
        for i in sheets:
            work = wb[i]

            for row in work.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)
            excel_playlist.append(excel_data)

        for item in excel_playlist:
            if ["Title", "Link", "Tag", "Progress"] in item:
                to_exclude = {0}
                data = [
                    element for i, element in enumerate(item) if i not in to_exclude
                ]
                for item in data:
                    try:
                        old_course = Course.objects.get(title=item[0])
                    except Course.DoesNotExist:
                        old_course = None
                    if old_course:
                        messages.info(
                            request, "course is already present in List of playlist"
                        )
                    else:
                        course = Course.objects.create(
                            author=request.user,
                            title=item[0],
                            link=item[1],
                            public=True,
                        )
                        course.tags.add(item[2])
                        course.save()
                    ydl = youtube_dl.YoutubeDL({"ignoreerrors": True, "quiet": True})
                    with ydl:
                        result = ydl.extract_info(
                            # We just want to extract the info
                            item[1],
                            download=False,
                        )

                    playlist_info = []
                    for video in result["entries"]:
                        video_url = dict(
                            (k, video[k])
                            for k in ["title", "duration", "webpage_url"]
                            if k in video
                        )
                        playlist_info.append(video_url.copy())
                    for i in playlist_info:
                        title = i.get("title")
                        duration = i.get("duration")
                        playlist_url = i.get("webpage_url")
                        list_store = PlaylistItem(
                            list_item=title,
                            time=duration,
                            playlist=result.get("title"),
                            link=playlist_url,
                        )
                        list_store.save()
            else:
                messages.info(
                    request,
                    "Format of excel sheet not meet as expected please Go through the sample excel file for correct format",
                )

            return redirect(Main)

    def get(self, request):
        return render(request, "course/upload.html")


class Download(View):
    """
    Download class is to download sample.xlsx file for correct excel file format.
    """

    def get(self, request):
        filename = "sample.xlsx"  # this is the file people must download
        with open(filename, "rb") as f:
            response = HttpResponse(f.read(), content_type="application/vnd.ms-excel")
            response["Content-Disposition"] = "attachment; filename=" + filename
            response["Content-Type"] = "application/vnd.ms-excel; charset=utf-16"
            return response
